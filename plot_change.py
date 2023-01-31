#!/usr/bin/env python
# coding: utf-8
# Natalia Sucher in the Kleen Lab, UCSF
# 1/31/2023

# plot_change.py: Generate significance and positive and negative mean differences of electrode weight changes at neuroanatomical sites

# OUTPUT: em_LL.xlsx, all_plot_change.xlsx, pos_plot_change.xlsx, neg_plot_change.xlsx

import time
import os
import pathlib
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import csv
import xlsxwriter


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

from scipy.io import loadmat
from scipy.signal import butter,filtfilt
from scipy.stats import ttest_1samp

neuroanat_list = ['frontalpole', #FRONTAL LOBE
    'parstriangularis',
    'parsopercularis',
    'parsorbitalis',
    'rostralmiddlefrontal',
    'caudalmiddlefrontal',
    'lateralorbitofrontal',
    'superiorfrontal',
    'medialorbitofrontal',
    'precentral',
    'postcentral', # PARIETAL LOBE
    'inferiorparietal',
    'superiorparietal',
    'supramarginal',
    'temporalpole', # TEMPORAL LOBE
    'middletemporal',
    'superiortemporal',
    'inferiortemporal',
    'parahippocampal',
    'Right-Hippocampus',
    'Left-Hippocampus',
    'Right-Amygdala',
    'Left-Amygdala',
    'entorhinal',
    'bankssts',
    'fusiform', # OCCIPITAL LOBE
    'lingual']
    # 'Right-Inf-Lat-Vent', # OTHER
    # 'Right-Cerebral-White-Matter',
    # 'Left-Cerebral-White-Matter',
    # 'Right-choroid-plexus',
    # 'Right-Putamen',
    # 'Right-VentralDC'];

abv_neuroanat_list = ['front-pole', #FRONTAL LOBE
    'parstri',
    'parsop',
    'parsorb',
    'rost-midfront',
    'caud-midfront',
    'latorb-front',
    'sup-front',
    'medorb-front',
    'precentral',
    'postcentral', # PARIETAL LOBE
    'inf-par',
    'sup-par',
    'supra-marg',
    'temp-pole', # TEMPORAL LOBE
    'mid-temp',
    'sup-temp',
    'inf-temp',
    'parahip',
    'R-Hip',
    'L-Hip',
    'R-Amyg',
    'L-Amyg',
    'entorhinal',
    'bankssts',
    'fusiform', # OCCIPITAL LOBE
    'lingual']
    # 'Right-Inf-Lat-Vent', # OTHER
    # 'Right-Cerebral-White-Matter',
    # 'Left-Cerebral-White-Matter',
    # 'Right-choroid-plexus',
    # 'Right-Putamen',
    # 'Right-VentralDC']


#    FUNCTIONS

def sem_w8s(LL,at_onset,before_onset,after_onset):
    row_num = np.shape(LL)[0]
    LL_meandiff = np.empty(row_num,)

    # AVERAGE ACTIVITY AFTER SYMPTOM - AVERAGE ACTIVITY BEFORE SYMPTOM
    for row in range(0,row_num-1): #for each channel in LL
         LL_meandiff[row] = np.mean(LL[row,int(at_onset):int(after_onset)]) - np.mean(LL[row,int(before_onset):int(at_onset)])
         # print(LL[row,int(before_onset):])
    return LL_meandiff
def ll_transform(llw,fs,d,bl_start,bl_stop):
    if bl_start == 0:
        sample_bl_start = 0
    elif bl_start > 0:
        sample_bl_start = (fs*bl_start)[0]
    sample_bl_end = (fs*bl_stop)[0]-1

    L = int(np.round(llw * fs) - 1)    # number of samples to calculate line length

    col_len = np.shape(d)[1]-L

    LL = np.empty([np.shape(d)[0],col_len])
    LL[:] = np.NaN

    for col_1 in range(0,col_len):
        LL[:,col_1] = np.sum(np.abs(np.diff(d[:,col_1:col_1+L])),1) # ~ 8 seconds timed

    for row_1 in range(0,np.shape(d)[0]):
        LL_nanmean = np.nanmean(LL[row_1, sample_bl_start:sample_bl_end])
        LL_nanstd = np.nanstd(LL[row_1, sample_bl_start:sample_bl_end])
        LL[row_1,:] = (LL[row_1,:] - LL_nanmean)/LL_nanstd

    return LL
def filt(d,fs):
    d_t = d.transpose()
    butter_array = np.array([1,(round(fs[0]/2)-1)])
    b,a = butter(2,butter_array/(fs[0]/2),btype='bandpass',output='ba')
    filt_d = filtfilt(b,a,d_t,axis=0,padtype='odd',padlen=3*(max(len(b),len(a))-1)).transpose()

    return filt_d
def load_elecs_anat(pt_path):

    #load anatomy and electrode matrix
    os.chdir(pt_path + 'Imaging/elecs')
    e_mat = loadmat('clinical_elecs_all.mat')
    anat_col = e_mat['anatomy'][:,3]

    return anat_col
def get_params(df_params,pt):
    params_bl_start = df_params.loc[pt]['BLstart']
    params_bl_stop = df_params.loc[pt]['BLstop']
    params_llw = df_params.loc[pt]['llw']

    return params_bl_start, params_bl_stop, params_llw
def open_xl(xl_name):
    df = pd.read_excel(xl_name, index_col=0, engine='openpyxl')

    return df
def good_ch(badch,anat,d,a_i):
    badch[a_i][0] = 1  # prepare for anatomy not in neuroanat list to be deleted in d and anat
    bad_logical = np.any(badch, 1)

    good_logical = ~bad_logical
    good_d = d[good_logical, :]
    good_anat = anat[good_logical]

    return good_logical, good_d, good_anat
def input_names(ptsz_input,sxmx_input,avg_path):
    pt_name, sz_name = ptsz_input.split('-')
    ptsz_name = pt_name + '_' + sz_name

    sx_name, mx_name = sxmx_input.split(' ')

    pt_path = avg_path + pt_name + '/'

    sz_path = pt_path + ptsz_name

    return pt_name, sz_name, sx_name, mx_name, ptsz_name, pt_path, sz_path
def file_suffix(ptsz_name):
    mat_name = ptsz_name + '.mat'  # mat of ppEEG data
    badch_name = ptsz_name + '_badch.mat'  # mat of bad channels
    csv_name = ptsz_name + '_mat.csv'  # csv of timecourse of sxmx every 200 ms

    return mat_name, badch_name, csv_name
def write_cell(cell_row, cell_col, cell_input, sheet_name):
    cell_location = sheet_name.cell(row=cell_row, column=cell_col)
    cell_location.value = cell_input
def sx_onset(fs, perdur_input, first_mx):
    time_ms = (first_mx / 5)

    at_onset = np.round(time_ms * fs)
    before_onset = np.round((time_ms - float(perdur_input)) * fs)
    after_onset = np.round((time_ms + float(perdur_input)) * fs)


    return at_onset, before_onset, after_onset, time_ms
def sz_mat(mat_name):
    sz_mat = loadmat(mat_name)  # load frame speed and ppEEG
    fs = sz_mat['fs'].flatten()
    d = sz_mat['ppEEG'][0:np.shape(anat)[0], :].astype(np.single)

    return fs, d
def badch_mat(badch_name):
    badch_mat = loadmat(badch_name)
    badch = badch_mat['bad_chs']
    badch = np.delete(badch, range(np.shape(anat)[0], np.shape(badch)[0]), 0)

    return badch
def create_workbooks(name_xlsx):
    if name_xlsx not in os.listdir():
        workbook_created = Workbook()
        sheet = workbook_created.active
        sheet.title = sxmx_input

        write_cell(1, 1, sxmx_input, sheet)
        write_cell(1, ptsz_i + 1, ptsz_name, sheet)

        workbook_created.save(opscea_data_path + name_xlsx)
        workbook_created.close()

        return name_xlsx
def write_new_sheet(sxmx_input, ptsz_i, ptsz_name, workbook_name):
    if sxmx_input not in workbook_name.sheetnames:
        workbook_name.create_sheet(title=sxmx_input, index=round(sxmx_count - 1))
        sheet_sxmx = workbook_name[sxmx_input]

        write_cell(1, 1, sxmx_input, sheet_sxmx)
        write_cell(1, ptsz_i + 1, ptsz_name, sheet_sxmx)

        workbook_name.save(opscea_data_path + name_xlsx)
        workbook_name.close()

print('help I am stuck in the code')
print(ptsz_input,' ',sxmx_input)

#   INITIALIZE NAMES
pt_name, sz_name, sx_name, mx_name, ptsz_name, pt_path, sz_path = input_names(ptsz_input,sxmx_input,avg_path)
list_str_xlsx = ['Neurosemiology.xlsx','pos_sign_change.xlsx','neg_sign_change.xlsx']

#   OPSCEA DATA PATH
os.chdir(avg_path + '/..')
opscea_data_path = os.getcwd() + '/'

#   CREATE NEW XLSX
for str_xlsx in range(0, len(list_str_xlsx)):
    create_workbooks(list_str_xlsx[str_xlsx])

workbook_loaded_pv = load_workbook(filename='Neurosemiology.xlsx')
workbook_loaded_pos = load_workbook(filename='pos_sign_change.xlsx')
workbook_loaded_neg = load_workbook(filename='neg_sign_change.xlsx')

workbook_list = [workbook_loaded_pv, workbook_loaded_pos, workbook_loaded_neg]

#   CREATE NEW SHEET OF SYMPTOM/MODE
for w_i in range(0, len(workbook_list)):
    workbook_name = workbook_list[w_i]

    # if sxmx_input not in workbook_loaded_pv.sheetnames:
    #     workbook_name.create_sheet(title=sxmx_input, index=round(sxmx_count - 1))
    if sxmx_input not in workbook_name.sheetnames:
        workbook_name.create_sheet(title=sxmx_input, index=round(sxmx_count - 1))

    sheet_sxmx = workbook_name[sxmx_input]


    write_cell(1, 1, sxmx_input, sheet_sxmx)
    write_cell(1, ptsz_i + 1, ptsz_name, sheet_sxmx)

df_params = open_xl('OPSCEAparams.xlsx')  # Ndimensions and params_list

#   PARAMS
bl_start, bl_stop, llw = get_params(df_params, pt_name) # 2 = params_llw
del df_params

#   GET ANATOMY LABELS
anat = load_elecs_anat(pt_path)

# PTSZ PATH
os.chdir(pt_path + ptsz_name + '/')

mat_name, badch_name, csv_name = file_suffix(ptsz_name)

fs, d = sz_mat(mat_name) #fs = frame speed; d = ppEEG

badch = badch_mat(badch_name)

sz_count = round(sz_count)

# OPSCEA DATA PATH
os.chdir(avg_path + '..')
opscea_data_path = os.getcwd() + '/'


# FILL IN XLSX WITH ROI LABELS
for w_i in range(0, len(workbook_list)):
    workbook_name = workbook_list[w_i]
    w_i_sheet_sxmx = workbook_name[sxmx_input]

    for n_l in range(0, len(neuroanat_list)):
        write_cell(n_l+2, 1, neuroanat_list[n_l], w_i_sheet_sxmx)

# CREATE GOOD CHANNEL LIST FOR LOGICAL, PPEEG (AKA D), AND ANATOMY
for a_i in range(0,np.shape(anat)[0]):

    if np.size(anat[a_i]) > 0:
        if anat[a_i][0] not in neuroanat_list:

            good_logical, good_d, good_anat = good_ch(badch, anat, d, a_i)

os.chdir(pt_path + ptsz_name + '/')

sx_csv = pd.read_csv(csv_name, usecols=[sx_name])
sx_vec = sx_csv.values.tolist()

pval = np.empty(np.shape(neuroanat_list))
pval[:] = np.NaN

if np.float64(mx_name) in sx_vec:
    filt_d = filt(good_d, fs)  # Filter out < 1 Hz (and up to nyquist)
    LL = ll_transform(llw, fs, filt_d, bl_start, bl_stop)   # ~8.5 seconds for EC96_01
    first_mx = sx_vec.index(np.float64(mx_name))
    at_onset, before_onset, after_onset, time_ms = sx_onset(fs, perdur_input, first_mx)
    LL_meandiff = sem_w8s(LL, at_onset, before_onset, after_onset)

    pval_list = np.empty(np.shape(neuroanat_list))
    pval_list[:] = np.NaN

    anat_index = [np.NaN] * np.shape(good_anat)[0]
    anat_list = np.empty(np.shape(good_anat)).tolist()

    for g in range(0, np.shape(good_anat)[0]):
        if len(good_anat[g]) > 0:
            anat_list[g] = good_anat[g][0]

    for n_l in range(0, len(neuroanat_list)):
        anat_w8s = []
        anat_w8s_label = []
        neg_change = []

        # print(neuroanat_list[n_l])
        for a_i in range(0, len(anat_list)):
            if neuroanat_list[n_l] == anat_list[a_i][:]:
                anat_w8s_label.append(neuroanat_list[n_l])
                anat_w8s.append(LL_meandiff[a_i])
                activity_change = float(LL_meandiff[a_i])
                if activity_change > 0:
                    pos_sheet_sxmx = workbook_loaded_pos[sxmx_input]
                    write_cell(n_l + 2, ptsz_i + 1, activity_change, pos_sheet_sxmx)
                elif activity_change < 0:
                    neg_sheet_sxmx = workbook_loaded_neg[sxmx_input]
                    write_cell(n_l + 2, ptsz_i + 1, activity_change, neg_sheet_sxmx)
        tstat, pval = ttest_1samp(anat_w8s, 0)
        sheet_sxmx = workbook_loaded_pv[sxmx_input]
        write_cell(n_l + 2, ptsz_i + 1, pval, sheet_sxmx)
else:
    pass

workbook_loaded_pv.save(opscea_data_path + 'Neurosemiology.xlsx')
workbook_loaded_pos.save(opscea_data_path + 'pos_sign_change.xlsx')
workbook_loaded_neg.save(opscea_data_path + 'neg_sign_change.xlsx')

workbook_loaded_pv.close()
workbook_loaded_pos.close()
workbook_loaded_neg.close()