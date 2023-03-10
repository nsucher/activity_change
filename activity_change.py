#!/usr/bin/env python
# coding: utf-8
# Natalia Sucher in the Kleen Lab, UCSF
# 1/31/2023

# plot_change.py: Generate significance and positive and negative mean differences of electrode weight changes at neuroanatomical sites

# OUTPUT:
#       neurosem_pvals.xlsx (pvals per ROI)
#       pos_plot_change.xlsx (positive electrical activity compared to the mean)
#       neg_plot_change.xlsx (negative ")

import time
import os
import pathlib
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import csv
import xlsxwriter


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

from scipy.io import loadmat
from scipy.signal import butter,filtfilt
from scipy.stats import ttest_1samp

neuroanat_list = ['frontalpole', #FRONTAL LOBE
    'parstriangularis',
    'parsopercularis',
    'parsorbitalis',
    'rostralmiddlefrontal',
    'caudalmiddlefrontal',
    'lateralorbitofrontal',
    'superiorfrontal',
    'medialorbitofrontal',
    'precentral',
    'postcentral', # PARIETAL LOBE
    'inferiorparietal',
    'superiorparietal',
    'supramarginal',
    'temporalpole', # TEMPORAL LOBE
    'middletemporal',
    'superiortemporal',
    'inferiortemporal',
    'parahippocampal',
    'Right-Hippocampus',
    'Left-Hippocampus',
    'Right-Amygdala',
    'Left-Amygdala',
    'entorhinal',
    'bankssts',
    'fusiform', # OCCIPITAL LOBE
    'lingual']
    # 'Right-Inf-Lat-Vent', # OTHER
    # 'Right-Cerebral-White-Matter',
    # 'Left-Cerebral-White-Matter',
    # 'Right-choroid-plexus',
    # 'Right-Putamen',
    # 'Right-VentralDC'];

abv_neuroanat_list = ['front-pole', #FRONTAL LOBE
    'parstri',
    'parsop',
    'parsorb',
    'rost-midfront',
    'caud-midfront',
    'latorb-front',
    'sup-front',
    'medorb-front',
    'precentral',
    'postcentral', # PARIETAL LOBE
    'inf-par',
    'sup-par',
    'supra-marg',
    'temp-pole', # TEMPORAL LOBE
    'mid-temp',
    'sup-temp',
    'inf-temp',
    'parahip',
    'R-Hip',
    'L-Hip',
    'R-Amyg',
    'L-Amyg',
    'entorhinal',
    'bankssts',
    'fusiform', # OCCIPITAL LOBE
    'lingual']
    # 'Right-Inf-Lat-Vent', # OTHER
    # 'Right-Cerebral-White-Matter',
    # 'Left-Cerebral-White-Matter',
    # 'Right-choroid-plexus',
    # 'Right-Putamen',
    # 'Right-VentralDC']


#    FUNCTIONS

def sem_w8s(LL,at_onset,before_onset,after_onset):
    row_num = np.shape(LL)[0]
    LL_meandiff = np.empty(row_num,)

    # AVERAGE ACTIVITY AFTER SYMPTOM - AVERAGE ACTIVITY BEFORE SYMPTOM
    for row in range(0,row_num-1): #for each channel in LL
         LL_meandiff[row] = np.mean(LL[row,int(at_onset):int(after_onset)]) - np.mean(LL[row,int(before_onset):int(at_onset)])
         # print(LL[row,int(before_onset):])
    return LL_meandiff
def ll_transform(llw,fs,d,bl_start,bl_stop):
    if bl_start == 0:
        sample_bl_start = 0
    elif bl_start > 0:
        sample_bl_start = (fs*bl_start)[0]
    sample_bl_end = (fs*bl_stop)[0]-1

    L = int(np.round(llw * fs) - 1)    # number of samples to calculate line length

    col_len = np.shape(d)[1]-L

    LL = np.empty([np.shape(d)[0],col_len])
    LL[:] = np.NaN

    for col_1 in range(0,col_len):
        LL[:,col_1] = np.sum(np.abs(np.diff(d[:,col_1:col_1+L])),1) # ~ 8 seconds timed

    for row_1 in range(0,np.shape(d)[0]):
        LL_nanmean = np.nanmean(LL[row_1, sample_bl_start:sample_bl_end])
        LL_nanstd = np.nanstd(LL[row_1, sample_bl_start:sample_bl_end])
        LL[row_1,:] = (LL[row_1,:] - LL_nanmean)/LL_nanstd

    return LL
def filt(d,fs):
    d_t = d.transpose()
    butter_array = np.array([1,(round(fs[0]/2)-1)])
    b,a = butter(2,butter_array/(fs[0]/2),btype='bandpass',output='ba')
    filt_d = filtfilt(b,a,d_t,axis=0,padtype='odd',padlen=3*(max(len(b),len(a))-1)).transpose()

    return filt_d
def load_elecs(pt_path):

    #load anatomy and electrode matrix
    os.chdir(pt_path + 'Imaging/elecs')
    e_mat = loadmat('clinical_elecs_all.mat')
    anat = e_mat['anatomy'][:,3]
    em = e_mat['elecmatrix']

    em_len = np.shape(em)[0]
    anat_len = np.shape(anat)[0]

    #delete excess rows in elecmatrix and anatomy
    if np.shape(em)[0] < np.shape(anat)[0]:
        anat = np.delete(anat, [em_len, anat_len-1])
    elif np.shape(em)[0] > np.shape(anat)[0]:
        em = np.delete(em, [anat_len,em_len-1])

    return anat, em
def get_params(df_params,pt):
    params_bl_start = df_params.loc[pt]['BLstart']
    params_bl_stop = df_params.loc[pt]['BLstop']
    params_llw = df_params.loc[pt]['llw']

    return params_bl_start, params_bl_stop, params_llw
def open_xl(xl_name):
    df = pd.read_excel(xl_name, index_col=0, engine='openpyxl')

    return df
def good_ch(badch, d, anat, em,a_i):
    badch[a_i][0] = 1  # prepare for anatomy not in neuroanat list to be deleted in d and anat
    bad_logical = np.any(badch, 1)

    good_logical = ~bad_logical
    good_d = d[good_logical, :]
    good_anat = anat[good_logical]
    good_em = em[good_logical,:]

    return good_logical, good_d, good_anat, good_em
def input_names(ptsz_input,sxmx_input,avg_path):
    pt_name, sz_name = ptsz_input.split('-')
    ptsz_name = pt_name + '_' + sz_name

    sx_name, mx_name = sxmx_input.split(' ')

    pt_path = avg_path + pt_name + '/'

    sz_path = pt_path + ptsz_name

    return pt_name, sz_name, sx_name, mx_name, ptsz_name, pt_path, sz_path
def file_suffix(ptsz_name):
    mat_name = ptsz_name + '.mat'  # mat of ppEEG data
    badch_name = ptsz_name + '_badch.mat'  # mat of bad channels
    csv_name = ptsz_name + '_mat.csv'  # csv of timecourse of sxmx every 200 ms

    return mat_name, badch_name, csv_name
def write_cell(cell_row, cell_col, cell_input, sheet_name):
    cell_location = sheet_name.cell(row=cell_row, column=cell_col)
    cell_location.value = cell_input
def sx_onset(fs, perdur_input, first_mx):
    time_ms = (first_mx / 5)

    at_onset = np.round(time_ms * fs)
    before_onset = np.round((time_ms - float(perdur_input)) * fs)
    after_onset = np.round((time_ms + float(perdur_input)) * fs)


    return at_onset, before_onset, after_onset, time_ms
def sz_mat(mat_name):
    sz_mat = loadmat(mat_name)  # load frame speed and ppEEG
    fs = sz_mat['fs'].flatten()
    d = sz_mat['ppEEG'][0:np.shape(anat)[0], :].astype(np.single)

    return fs, d
def badch_mat(badch_name):
    badch_mat = loadmat(badch_name)
    badch = badch_mat['bad_chs']
    badch = np.delete(badch, range(np.shape(anat)[0], np.shape(badch)[0]), 0)

    return badch
def create_workbooks(name_xlsx):
    if name_xlsx not in os.listdir():
        workbook_created = Workbook()
        sheet = workbook_created.active
        sheet.title = sxmx_input

        write_cell(1, 1, sxmx_input, sheet)
        write_cell(1, ptsz_i + 1, ptsz_name, sheet)

        workbook_created.save(opscea_data_path + name_xlsx)
        workbook_created.close()
def create_em_workbook(name_xlsx):
    if name_xlsx not in os.listdir():
        workbook_created = Workbook()
        sheet = workbook_created.active

        write_cell(1, ptsz_i + 1, ptsz_name, sheet)

        workbook_created.save(opscea_data_path + name_xlsx)
        workbook_created.close()
def write_new_sheet(sxmx_input, ptsz_i, ptsz_name, workbook_name):
    if sxmx_input not in workbook_name.sheetnames:
        workbook_name.create_sheet(title=sxmx_input, index=round(sxmx_count - 1))
        sheet_sxmx = workbook_name[sxmx_input]

        write_cell(1, 1, sxmx_input, sheet_sxmx)
        write_cell(1, ptsz_i + 1, ptsz_name, sheet_sxmx)

        workbook_name.save(opscea_data_path + name_xlsx)
        workbook_name.close()
def find_laterality(em_sign_row):
    isR = np.nansum(em_sign_row) > 0
    isL = isR != 1

    if isR:
        laterality = 'r'
    elif isL:
        laterality = 'l'

    return laterality


print('help I am stuck in the code')
print(ptsz_input,' ',sxmx_input)

#   INITIALIZE NAMES
pt_name, sz_name, sx_name, mx_name, ptsz_name, pt_path, sz_path = input_names(ptsz_input,sxmx_input,avg_path)
list_str_xlsx = ['neurosem_pvals.xlsx', 'all_sign_change.xlsx', 'pos_sign_change.xlsx', 'neg_sign_change.xlsx', 'elec_w8s.xlsx']

#   OPSCEA DATA PATH
os.chdir(avg_path + '/..')
opscea_data_path = os.getcwd() + '/'

#   CREATE NEW XLSX WORKBOOKS
for str_xlsx in range(0, len(list_str_xlsx)):
    create_workbooks(list_str_xlsx[str_xlsx])
create_em_workbook('elec_matrix.xlsx')


workbook_loaded_pv = load_workbook(filename='neurosem_pvals.xlsx')
workbook_loaded_all = load_workbook(filename='all_sign_change.xlsx')
workbook_loaded_pos = load_workbook(filename='pos_sign_change.xlsx')
workbook_loaded_neg = load_workbook(filename='neg_sign_change.xlsx')
workbook_loaded_em = load_workbook(filename='elec_matrix.xlsx')
workbook_loaded_w8s = load_workbook(filename='elec_w8s.xlsx')

workbook_list = [workbook_loaded_pv, workbook_loaded_all, workbook_loaded_pos, workbook_loaded_neg, workbook_loaded_w8s] # elec matrix not included

#   CREATE NEW SHEET OF SYMPTOM/MODE
for w_i in range(0, len(workbook_list)):
    workbook_name = workbook_list[w_i]
    if sxmx_input not in workbook_name.sheetnames:
        workbook_name.create_sheet(title=sxmx_input, index=round(sxmx_count - 1))
    sheet_sxmx = workbook_name[sxmx_input]
    write_cell(1, 1, sxmx_input, sheet_sxmx)
    write_cell(1, ptsz_i + 1, ptsz_name, sheet_sxmx)

df_params = open_xl('OPSCEAparams.xlsx')  # Ndimensions and params_list

#   PARAMS
bl_start, bl_stop, llw = get_params(df_params, pt_name) # 2 = params_llw
del df_params

#   GET ANATOMY LABELS
anat, em = load_elecs(pt_path)

sheet_em = workbook_loaded_em.active

#       SEIZURE NAMES
for each_col1 in range(0, np.shape(em)[1]):
    write_cell(1, 3 * ptsz_i - 1 + each_col1, ptsz_name, sheet_em)

# szxyz = []
#       3D COORDINATES
# for em_row in range(0, np.shape(em)[0]):
#     for em_col in range(0, np.shape(em)[1]):
#         write_cell(em_row + 2, 3 * ptsz_i - 1 + em_col, em[em_row][em_col], sheet_em)
#
#
# em_sign_row = em[:,]
# laterality = find_laterality(em_sign_row)


#   PTSZ PATH
os.chdir(pt_path + ptsz_name + '/')

#   LOAD MAT FILES
mat_name, badch_name, csv_name = file_suffix(ptsz_name)
fs, d = sz_mat(mat_name) #fs = frame speed; d = ppEEG
badch = badch_mat(badch_name)
sz_count = round(sz_count)

# OPSCEA DATA PATH
os.chdir(avg_path + '..')
opscea_data_path = os.getcwd() + '/'

# FILL IN XLSX WITH ROI LABELS
# for w_i in range(0, len(workbook_list)):
#     workbook_name = workbook_list[w_i]
#     w_i_sheet_sxmx = workbook_name[sxmx_input]
#
#     for n_l in range(0, len(neuroanat_list)):
#         write_cell(n_l+2, 1, neuroanat_list[n_l], w_i_sheet_sxmx)


pv_sheet = workbook_loaded_pv[sxmx_input]
for n_l in range(0, len(neuroanat_list)):
    write_cell(n_l+2, 1, neuroanat_list[n_l], pv_sheet)


# CREATE GOOD CHANNEL LIST FOR LOGICAL, PPEEG (AKA D), AND ANATOMY
for a_i in range(0,np.shape(anat)[0]):

    if np.size(anat[a_i]) > 0:
        if anat[a_i][0] not in neuroanat_list:

            good_logical, good_d, good_anat, good_em = good_ch(badch, d, anat, em, a_i)

#   FILL IN  XLSX OF 3D ELEC MATRIX
sheet_em = workbook_loaded_em.active

# print(good_anat)
#       3D COORDINATES
for em_row in range(0, np.shape(good_em)[0]):
    # print(good_anat[em_row][0])
    write_cell(em_row + 2, 1, good_anat[em_row][0], sheet_em)
    for em_col in range(0, np.shape(good_em)[1]):
        write_cell(em_row + 2, 3 * ptsz_i - 1 + em_col, good_em[em_row][em_col], sheet_em)


em_sign_row = good_em[:,]
laterality = find_laterality(em_sign_row)

os.chdir(pt_path + ptsz_name + '/')

sx_csv = pd.read_csv(csv_name, usecols=[sx_name])
sx_vec = sx_csv.values.tolist()

pval = np.empty(np.shape(neuroanat_list))
pval[:] = np.NaN

if np.float64(mx_name) in sx_vec:
    filt_d = filt(good_d, fs)  # Filter out < 1 Hz (and up to nyquist)
    LL = ll_transform(llw, fs, filt_d, bl_start, bl_stop)   # ~8.5 seconds for EC96_01
    first_mx = sx_vec.index(np.float64(mx_name))
    at_onset, before_onset, after_onset, time_ms = sx_onset(fs, perdur_input, first_mx)
    LL_meandiff = sem_w8s(LL, at_onset, before_onset, after_onset)

    pval_list = np.empty(np.shape(neuroanat_list))
    pval_list[:] = np.NaN

    anat_index = [np.NaN] * np.shape(good_anat)[0]
    anat_list = np.empty(np.shape(good_anat)).tolist()

    for g in range(0, np.shape(good_anat)[0]):
        if len(good_anat[g]) > 0:
            anat_list[g] = good_anat[g][0]

    for electrode in range(0,np.shape(LL_meandiff)[0]):
        activity_change = float(LL_meandiff[electrode])
        all_sheet_sxmx = workbook_loaded_all[sxmx_input]
        if activity_change < 400 and activity_change > -400:
            write_cell(electrode + 2, ptsz_i + 1, activity_change, all_sheet_sxmx)
            if activity_change > 0:
                pos_sheet_sxmx = workbook_loaded_pos[sxmx_input]
                write_cell(electrode + 2, ptsz_i + 1, activity_change, pos_sheet_sxmx)
            elif activity_change < 0:
                neg_sheet_sxmx = workbook_loaded_neg[sxmx_input]
                write_cell(electrode + 2, ptsz_i + 1, activity_change, neg_sheet_sxmx)

    w8s_array = []
    anat_array = []
    for n_l in range(0, len(neuroanat_list)):
        anat_w8s = []
        anat_w8s_label = []

        # sheet_w8s = workbook_loaded_w8s[sxmx_input]
        # print(neuroanat_list[n_l])
        for a_i in range(0, len(anat_list)):
            if neuroanat_list[n_l] == anat_list[a_i][:]:
                if LL_meandiff[a_i] < 400:
                    if LL_meandiff[a_i] > -400:
                        anat_w8s_label.append(neuroanat_list[n_l])
                        anat_w8s.append(LL_meandiff[a_i])
                        sheet_w8s = workbook_loaded_w8s[sxmx_input]
                        write_cell(a_i + 2, 1, anat_w8s_label[-1], sheet_w8s)
                        write_cell(a_i + 2, ptsz_i+1, anat_w8s[-1], sheet_w8s)
                # activity_change = float(LL_meandiff[a_i])
                # all_sheet_sxmx = workbook_loaded_all[sxmx_input]
        sheet_w8s = workbook_loaded_w8s[sxmx_input]
        w8s_array.append(anat_w8s)
        anat_array.append(anat_w8s_label)
        # write_cell(a_i + 2, ptsz_i + 1, w8s_array[-1], sheet_w8s)
        tstat, pval = ttest_1samp(anat_w8s, 0)
        sheet_sxmx = workbook_loaded_pv[sxmx_input]
        write_cell(n_l + 2, ptsz_i + 1, pval, sheet_sxmx)

else:
    pass


workbook_loaded_pv.save(opscea_data_path + 'neurosem_pvals.xlsx')
workbook_loaded_all.save(opscea_data_path + 'all_sign_change.xlsx')
workbook_loaded_pos.save(opscea_data_path + 'pos_sign_change.xlsx')
workbook_loaded_neg.save(opscea_data_path + 'neg_sign_change.xlsx')
workbook_loaded_em.save(opscea_data_path + 'elec_matrix.xlsx')
workbook_loaded_w8s.save(opscea_data_path + 'elec_w8s.xlsx')

workbook_loaded_pv.close()
workbook_loaded_all.close()
workbook_loaded_pos.close()
workbook_loaded_neg.close()
workbook_loaded_em.close()
workbook_loaded_w8s.close()